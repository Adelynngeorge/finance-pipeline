import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

debts = pd.read_csv('debts_clean.csv')
sims  = pd.read_csv('simulations.csv')

summary = pd.DataFrame({
    'Metric': ['Total Debt', 'Monthly Payments', 'Monthly Interest', 'Number of Accounts'],
    'Value':  [
        f"${debts['balance'].sum():,.2f}",
        f"${debts['min_payment'].sum():,.2f}",
        f"${debts['monthly_interest'].sum():,.2f}",
        str(len(debts))
    ]
})

with pd.ExcelWriter('finance_dashboard.xlsx', engine='openpyxl') as writer:
    debts.to_excel(writer,   sheet_name='Debts',      index=False)
    sims.to_excel(writer,    sheet_name='Simulations', index=False)
    summary.to_excel(writer, sheet_name='Summary',     index=False)

wb = load_workbook('finance_dashboard.xlsx')
header_fill = PatternFill("solid", fgColor="1D9E75")
header_font = Font(color="FFFFFF", bold=True)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 20

wb.save('finance_dashboard.xlsx')
print("Done! finance_dashboard.xlsx created.")